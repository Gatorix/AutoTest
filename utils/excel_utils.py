from openpyxl import load_workbook
import xlrd

from utils.file_utils import remove_file


def diff_between_xlsx(ori_xlsx, latest_xlsx):
    ori_wb = load_workbook(ori_xlsx)
    ori_sheet = ori_wb[ori_wb.sheetnames[0]]
    latest_wb = load_workbook(latest_xlsx)
    latest_sheet = latest_wb[latest_wb.sheetnames[0]]
    diffs = []
    for row_idx, row in enumerate(ori_sheet.values):
        try:
            for value_idx, value in enumerate(row):
                if value != latest_sheet.cell(row_idx + 1, value_idx + 1).value:
                    cell_num = str(latest_sheet.cell(row_idx + 1, value_idx + 1)).split(".")[-1].replace(">", "")
                    latest_value = latest_sheet.cell(row_idx + 1, value_idx + 1).value
                    diffs.append(f'单元格{cell_num}数据不一致：预置文件为"{value}"，下载文件为"{latest_value}"')
        except IndexError:
            diffs.append(f'excel文件行数不一致：预置文件{ori_sheet.max_row}行，下载文件{latest_sheet.max_row}行')
            break
    remove_file(latest_xlsx)
    if diffs:
        raise Exception('\n'.join(diffs))
    else:
        return True


def diff_between_xls(ori_xls, latest_xls):
    ori_book = xlrd.open_workbook(ori_xls)
    ori_sheet = ori_book.sheet_by_index(0)
    latest_book = xlrd.open_workbook(latest_xls)
    latest_sheet = latest_book.sheet_by_index(0)
    diffs = []
    for row_idx in range(ori_sheet.nrows):
        try:
            for col_idx, value in enumerate(ori_sheet.row_values(row_idx)):
                if value != latest_sheet.cell(row_idx, col_idx).value:
                    latest_value = latest_sheet.cell(row_idx, col_idx).value
                    diffs.append(
                        f'第{row_idx + 1}行-第{col_idx + 1}列数据不一致：预置文件为"{value}"，下载文件为"{latest_value}"')
        except IndexError:
            diffs.append(f'excel文件行数不一致：预置文件{ori_sheet.nrows}行，下载文件{latest_sheet.nrows}行')
            break
    remove_file(latest_xls)
    if diffs:
        raise Exception('对比结果：\n'+'\n'.join(diffs))
    else:
        return True


def check_excel_diff(ori, latest):
    if all([ori.split('.')[-1] == 'xls', latest.split('.')[-1] == 'xls']):
        return diff_between_xls(ori, latest)
    elif all([ori.split('.')[-1] == 'xlsx', latest.split('.')[-1] == 'xlsx']):
        return diff_between_xlsx(ori, latest)
    else:
        raise Exception(f'文件类型不一致，无法对比！\n{ori}\n{latest}')
